"""
导出功能集成测试
"""
import os
import sys
import json
import pytest
import tempfile
import io

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'server'))


@pytest.fixture(autouse=True)
def setup_db(monkeypatch):
    import database
    db_dir = os.path.join(tempfile.gettempdir(), 'attendance-v3-test-export')
    db_path = os.path.join(db_dir, 'attendance.db')
    monkeypatch.setattr(database, 'DB_DIR', db_dir)
    monkeypatch.setattr(database, 'DB_PATH', db_path)
    if os.path.exists(db_path):
        os.remove(db_path)
    database.init_db()
    yield db_path
    try:
        os.remove(db_path)
    except OSError:
        pass


class MockHandler:
    def __init__(self, body=None, path=''):
        self.sent = None
        self._body = json.dumps(body) if body else None
        self.path = path
        body_b = json.dumps(body).encode() if body else b''
        self.headers = {'Content-Length': str(len(body_b))}
        self.rfile = type('obj', (object,), {'read': lambda s, n: body_b})()

    def _read_body(self):
        if self._body:
            return json.loads(self._body)
        return {}

    def _send_json(self, code, data=None, message=None):
        self.sent = {'code': code, 'data': data, 'message': message}

    def send_response(self, code):
        self.sent = {'_status': code}

    def send_header(self, k, v):
        pass

    def end_headers(self):
        pass

    def wfile(self):
        return type('obj', (object,), {'write': lambda s, d: None})()


class TestExport:
    def test_build_flat_report_creates_workbook(self):
        from handlers.export import build_flat_report
        import openpyxl
        import io

        records = [
            {'employeeNo': 'T001', 'name': '张三', 'department': '技术部',
             'date': '2026-07-01', 'signIn': '08:30', 'signOut': '17:30', 'status': 'normal'}
        ]
        template = {'fields': [
            {'field': 'employeeNo', 'label': '工号'},
            {'field': 'name', 'label': '姓名'},
            {'field': 'signIn', 'label': '签到'},
            {'field': 'signOut', 'label': '签退'},
            {'field': 'status', 'label': '状态'},
        ]}

        output = build_flat_report(records, template, 'test.xlsx')
        wb = openpyxl.load_workbook(output, data_only=True)
        ws = wb.active
        assert ws.max_row == 2
        assert ws.max_column == 5
        assert ws.cell(1, 1).value == '工号'
        assert ws.cell(2, 1).value == 'T001'
        assert ws.cell(2, 5).value == '正常'

    def test_build_flat_report_conditional_formatting(self):
        from handlers.export import build_flat_report
        import openpyxl
        import io

        records = [
            {'employeeNo': 'T001', 'name': '张三', 'signIn': '09:30', 'signOut': '16:30', 'status': 'normal'}
        ]
        template = {'fields': [
            {'field': 'signIn', 'label': '签到'},
            {'field': 'signOut', 'label': '签退'},
        ]}

        output = build_flat_report(records, template, 'test.xlsx', startTime='08:30', endTime='17:30')
        wb = openpyxl.load_workbook(output)
        ws = wb.active

        cf_count = len(ws.conditional_formatting._cf_rules)
        assert cf_count >= 2

    def test_build_calendar_report_structure(self):
        from handlers.export import build_calendar_report
        import openpyxl

        results = [
            {'employeeNo': 'T001', 'name': '张三', 'department': '技术部',
             'date': '2026-07-01', 'status': 'normal', 'signIn': '08:30', 'signOut': '17:30'}
        ]
        output = build_calendar_report('2026-07', [], results, [])
        wb = openpyxl.load_workbook(output, data_only=True)
        ws = wb.active
        assert ws.title == '2026年7月考勤明细'
        assert ws.max_row >= 2

    def test_build_calendar_with_holiday(self):
        from handlers.export import build_calendar_report
        import openpyxl

        results = [
            {'employeeNo': 'T001', 'name': '张三', 'department': '技术部',
             'date': '2026-07-01', 'status': 'normal', 'signIn': '08:30', 'signOut': '17:30'}
        ]
        holidays = [{'date': '2026-07-01', 'name': '法定假日', 'is_holiday': True}]
        output = build_calendar_report('2026-07', [], results, [], holidays)
        wb = openpyxl.load_workbook(output, data_only=True)
        ws = wb.active
        assert ws.cell(3, 2).value == '法定假日'

    def test_build_calendar_groups_by_department(self):
        from handlers.export import build_calendar_report
        import openpyxl

        results = [
            {'employeeNo': 'T001', 'name': '张三', 'department': '技术部',
             'date': '2026-07-01', 'status': 'normal', 'signIn': '08:30', 'signOut': '17:30'},
            {'employeeNo': 'T002', 'name': '李四', 'department': '技术部',
             'date': '2026-07-01', 'status': 'normal', 'signIn': '08:30', 'signOut': '17:30'},
            {'employeeNo': 'S001', 'name': '王五', 'department': '销售部',
             'date': '2026-07-01', 'status': 'leave', 'leaveType': '年假', 'leaveHours': 8},
        ]
        output = build_calendar_report('2026-07', [], results, [])
        wb = openpyxl.load_workbook(output, data_only=True)
        ws = wb.active

        header1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        header2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]

        # 部门名只在部门首列出现（合并单元格），技术部两人连续
        assert header1.count('技术部') == 1
        assert header1.count('销售部') == 1
        assert '张三' in header2 and '李四' in header2 and '王五' in header2

        tech_idx = header1.index('技术部')
        assert header1[tech_idx + 1] in ('', None), 'second tech employee keeps dept header empty'
        assert header1[tech_idx + 2] == '销售部', 'sales dept follows tech block'
