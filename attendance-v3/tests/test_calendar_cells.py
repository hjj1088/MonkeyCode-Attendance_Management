"""
Calendar export cell content tests - V3.1.1
Validates build_am_cell / build_pm_cell priority rules:
  leave > signIn > travel > missTime > absent/rest
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'server'))

import openpyxl
import io
from handlers.export import build_calendar_report


def _make_result(eno, date, **kw):
    base = {
        'employeeNo': eno, 'name': '张三', 'department': '研发部',
        'date': date, 'signIn': '', 'signOut': '', 'status': 'normal',
        'leaveType': '', 'leaveHours': 0, 'travelHours': 0, 'absent': False,
        'missTime': '', 'sourceTravelIds': [], 'sourceMissIds': [],
        'scheduleStart': '08:30', 'scheduleEnd': '17:30',
    }
    base.update(kw)
    return base


def _get_cells(results):
    import openpyxl
    output = build_calendar_report('2026-02', None, results, [], [], '08:30', '17:30')
    # output is BytesIO, load it
    wb = openpyxl.load_workbook(output, data_only=True)
    ws = wb.active
    d = int(results[0]['date'][8:10])
    row = 3 + (d - 1) * 2
    am = ws.cell(row, 4).value or ''
    pm = ws.cell(row + 1, 4).value or ''
    return am, pm


def test_travel_date_with_punch():
    """Travel with punch → punch side shows time, other side shows travel"""
    results = [
        _make_result('001', '2026-02-02', signIn='08:25', status='normal', travelHours=8, sourceTravelIds=[2])
    ]
    am, pm = _get_cells(results)
    assert am == '08:25', f"'{am}' != '08:25'"
    assert pm == '出差/8h', f"'{pm}' != '出差/8h'"


def test_travel_date_without_punch():
    """Travel without punch → both sides show travel"""
    results = [
        _make_result('001', '2026-02-02', status='travel', travelHours=8, sourceTravelIds=[2])
    ]
    am, pm = _get_cells(results)
    assert am == '出差/8h', f"'{am}' != '出差/8h'"
    assert pm == '出差/8h', f"'{pm}' != '出差/8h'"


def test_travel_punch_in_only():
    """Travel with only punch-in → AM=punch, PM=travel (missing side)"""
    results = [
        _make_result('001', '2026-02-02', signIn='08:25', status='normal', travelHours=8, sourceTravelIds=[2])
    ]
    am, pm = _get_cells(results)
    assert am == '08:25'
    assert pm == '出差/8h'


def test_travel_punch_out_only():
    """Travel with only punch-out → AM=travel (missing side), PM=punch"""
    results = [
        _make_result('001', '2026-02-02', signOut='17:35', status='normal', travelHours=8, sourceTravelIds=[2])
    ]
    am, pm = _get_cells(results)
    assert am == '出差/8h', f"'{am}' != '出差/8h'"
    assert pm == '17:35', f"'{pm}' != '17:35'"


def test_leave_am_and_pm_filled():
    """Leave → both AM/PM show leave type/hrs (no empty PM)"""
    results = [
        _make_result('001', '2026-02-02', status='leave', leaveType='事假', leaveHours=8)
    ]
    am, pm = _get_cells(results)
    assert am == '事假/8h', f"'{am}' != '事假/8h'"
    assert pm == '事假/8h', f"'{pm}' != '事假/8h'"


def test_absent_am_and_pm_filled():
    """Absent → both AM/PM show 缺勤"""
    results = [
        _make_result('001', '2026-02-02', status='absent', absent=True)
    ]
    am, pm = _get_cells(results)
    assert am == '缺勤', f"'{am}' != '缺勤'"
    assert pm == '缺勤', f"'{pm}' != '缺勤'"


def test_miss_punch_am():
    """Miss punch in AM → AM shows 漏打卡，PM shows punch if exists"""
    results = [
        _make_result('001', '2026-02-02', signOut='17:35', status='normal', missTime='上午', sourceMissIds=[1])
    ]
    am, pm = _get_cells(results)
    assert am == '漏打卡', f"'{am}' != '漏打卡'"
    assert pm == '17:35', f"'{pm}' != '17:35'"


def test_miss_punch_pm():
    """Miss punch in PM → AM shows punch, PM shows 漏打卡"""
    results = [
        _make_result('001', '2026-02-02', signIn='08:25', status='normal', missTime='下午', sourceMissIds=[1])
    ]
    am, pm = _get_cells(results)
    assert am == '08:25', f"'{am}' != '08:25'"
    assert pm == '漏打卡', f"'{pm}' != '漏打卡'"


def test_legacy_no_sign_in_shows_absent():
    """Legacy status=no_sign_in (with travel) → shows punch time, then travel (user rule)"""
    # With travel: travel/punch priority applies
    results = [
        _make_result('001', '2026-02-02', signOut='17:35', status='no_sign_in', travelHours=8, sourceTravelIds=[2])
    ]
    am, pm = _get_cells(results)
    # AM: no punch, travel present → travel; PM: has punch
    assert am == '出差/8h', f"'{am}' != '出差/8h'"
    assert pm == '17:35', f"'{pm}' != '17:35'"


def test_legacy_no_sign_out_shows_absent():
    """Legacy status=no_sign_out → PM shows 缺勤 if no travel"""
    results = [
        _make_result('001', '2026-02-02', signIn='08:25', status='no_sign_out')
    ]
    am, pm = _get_cells(results)
    # AM: has punch; PM: no punch, no travel → absent=缺勤
    assert am == '08:25', f"'{am}' != '08:25'"
    assert pm == '缺勤', f"'{pm}' != '缺勤'"


def test_rest_day_empty():
    """Rest day → both AM/PM empty"""
    results = [
        _make_result('001', '2026-02-02', status='rest')
    ]
    am, pm = _get_cells(results)
    assert am == '', f"'{am}' != ''"
    assert pm == '', f"'{pm}' != ''"


def test_travel_plus_leave_priority():
    """Travel + leave same day → leave takes priority (user spec: 请假 before 出差)"""
    results = [
        _make_result('001', '2026-02-02', status='travel', leaveType='调休', leaveHours=2, travelHours=8, sourceTravelIds=[2], sourceLeaveIds=[1])
    ]
    am, pm = _get_cells(results)
    # Leave wins per user spec
    assert am == '调休/2h', f"'{am}' != '调休/2h'"
    assert pm == '调休/2h', f"'{pm}' != '调休/2h'"


if __name__ == '__main__':
    import sys
    failures = 0
    for name, fn in list(globals().items()):
        if name.startswith('test_'):
            try:
                fn()
                print(f'PASS {name}')
            except AssertionError as e:
                print(f'FAIL {name}: {e}')
                failures += 1
    print()
    sys.exit(0 if failures == 0 else failures)
