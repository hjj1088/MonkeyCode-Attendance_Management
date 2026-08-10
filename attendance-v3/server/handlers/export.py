# V3.1 handlers/export.py
# Export handler - generates XLSX from POST body data
# Fully ported from V2.0 export_server.py

import json
import io
import re
from datetime import date
from urllib.parse import quote


def _send_xlsx(handler, output, filename):
    safe_filename = quote(filename)
    handler.send_response(200)
    handler.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    handler.send_header('Content-Disposition', f"attachment; filename*=UTF-8''{safe_filename}")
    handler.send_header('Access-Control-Allow-Origin', '*')
    handler.end_headers()
    handler.wfile.write(output.read())

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

RED_FONT = Font(color='FF0000')
BLUE_FONT = Font(color='0066CC')
GRAY_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
TEXT_FMT = '@'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def _get_cell_style(val):
    font = None
    fill = None
    if not val:
        return font, fill
    sv = str(val)
    if re.search(r'假|休|出差|加班|补卡', sv):
        font = BLUE_FONT
    elif re.search(r'迟|早|上班未打卡|下班未打卡|缺勤|漏打卡', sv):
        font = RED_FONT
    return font, fill


def _is_time_val(v):
    return bool(v and re.match(r'^\d{1,2}:\d{2}$', str(v)))


def _time_to_minutes(t):
    if not t or not _is_time_val(str(t)):
        return None
    parts = str(t).split(':')
    return int(parts[0]) * 60 + int(parts[1])


def build_flat_report(records, template, filename, startTime=None, endTime=None):
    """构建平铺报表 XLSX - 完全照搬 V2.0"""
    wb = Workbook()
    ws = wb.active
    ws.title = '考勤记录'

    headers = [f['label'] for f in template['fields']]

    status_labels = {
        'normal': '正常', 'rest': '休息', 'abnormal': '迟到',
        'leave': '请假', 'travel': '出差', 'absent': '缺勤',
        'overtime': '疑似加班', 'suspect_ot': '疑似加班',
        'no_sign_in': '缺勤', 'no_sign_out': '缺勤',
    }

    for c, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = Font(bold=True)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for i, rec in enumerate(records):
        row_idx = i + 2
        for ci, f in enumerate(template['fields']):
            col_idx = ci + 1
            field_name = f['field']

            if field_name == '_index':
                val = i + 1
            else:
                val = rec.get(field_name, '')
                if field_name == 'status':
                    val = status_labels.get(val, val)

            val_str = str(val) if val is not None else ''
            cell = ws.cell(row=row_idx, column=col_idx, value=val_str)
            cell.border = THIN_BORDER
            cell.number_format = TEXT_FMT

            font, fill = _get_cell_style(val_str)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill

            sst = rec.get('scheduleStart', '') or startTime
            sed = rec.get('scheduleEnd', '') or endTime
            if field_name == 'signIn' and val and _is_time_val(str(val)) and _is_time_val(sst):
                if _time_to_minutes(str(val)) > _time_to_minutes(sst):
                    cell.font = RED_FONT
            elif field_name == 'signOut' and val and _is_time_val(str(val)) and _is_time_val(sed):
                if _time_to_minutes(str(val)) < _time_to_minutes(sed):
                    cell.font = RED_FONT

    _st = startTime
    _et = endTime
    if not _st or not _et:
        for rec in records:
            sst = rec.get('scheduleStart', '')
            sed2 = rec.get('scheduleEnd', '')
            if sst and not _st:
                _st = sst
            if sed2 and not _et:
                _et = sed2
            if _st and _et:
                break
        if not _st or not _is_time_val(_st):
            _st = '08:30'
        if not _et or not _is_time_val(_et):
            _et = '17:30'

    sh = None; sm = None; eh = None; em = None
    if _st and _et and _is_time_val(_st) and _is_time_val(_et):
        sh, sm = _st.split(':')
        eh, em = _et.split(':')

    if sh is not None:
        data_end_row = len(records) + 1
        cf_count = 0
        for ci, f in enumerate(template['fields']):
            col_letter = get_column_letter(ci + 1)
            col_range = f'{col_letter}2:{col_letter}{data_end_row}'
            cell_ref = f'{col_letter}2'
            field = f.get('field', '')
            flabel = f.get('label', '')
            is_si = field == 'signIn' or ('签到' in flabel)
            is_so = field == 'signOut' or ('签退' in flabel)
            if is_si:
                late_rule = FormulaRule(
                    formula=[f'AND({cell_ref}<>"",NOT(ISERROR(TIMEVALUE({cell_ref}))),TIMEVALUE({cell_ref})>TIME({int(sh)},{int(sm)},0))'],
                    font=RED_FONT
                )
                ws.conditional_formatting.add(col_range, late_rule)
                cf_count += 1
            elif is_so:
                early_rule = FormulaRule(
                    formula=[f'AND({cell_ref}<>"",NOT(ISERROR(TIMEVALUE({cell_ref}))),TIMEVALUE({cell_ref})<TIME({int(eh)},{int(em)},0))'],
                    font=RED_FONT
                )
                ws.conditional_formatting.add(col_range, early_rule)
                cf_count += 1

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_width = len(headers[col_idx - 1]) * 2 + 4
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                if cell.value:
                    val_str = str(cell.value)
                    width = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    max_width = max(max_width, width + 2)
        ws.column_dimensions[col_letter].width = min(max_width, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_calendar_report(target_month, fields, results, schedules, holidays=None, startTime=None, endTime=None):
    """构建日历报表 XLSX - 完全照搬 V2.0"""
    y, m = map(int, target_month.split('-'))
    from calendar import monthrange
    _, last_day = monthrange(y, m)

    field_set = set(f['field'] for f in fields) if fields else set()
    def use_field(name):
        return not fields or name in field_set

    def build_am_cell(r):
        if not r:
            return ''
        if r.get('status') == 'rest':
            return ''
        if r.get('status') == 'leave' or r.get('leaveType'):
            parts = []
            if use_field('leaveType') and r.get('leaveType'):
                parts.append(r['leaveType'])
            if use_field('leaveHours') and r.get('leaveHours') is not None:
                parts.append(str(r['leaveHours']) + 'h')
            return '/'.join(parts) if parts else '请假'
        if use_field('signIn') and r.get('signIn'):
            return r['signIn']
        if r.get('sourceTravelIds') and r.get('travelHours'):
            parts = ['出差']
            if use_field('travelHours') and r.get('travelHours'):
                parts.append(str(r['travelHours']) + 'h')
            return '/'.join(parts)
        mt = str(r.get('missTime', ''))
        if r.get('missTime') and ('上午' in mt or '上班' in mt or 'AM' in mt):
            return '漏打卡'
        if r.get('status') in ('absent', 'no_sign_in', 'no_sign_out') or r.get('absent'):
            return '缺勤'
        return ''

    def build_pm_cell(r):
        if not r:
            return ''
        if r.get('status') == 'rest':
            return ''
        if r.get('status') == 'leave' or r.get('leaveType'):
            parts = []
            if use_field('leaveType') and r.get('leaveType'):
                parts.append(r['leaveType'])
            if use_field('leaveHours') and r.get('leaveHours') is not None:
                parts.append(str(r['leaveHours']) + 'h')
            return '/'.join(parts) if parts else '请假'
        if use_field('signOut') and r.get('signOut'):
            return r['signOut']
        if r.get('sourceTravelIds') and r.get('travelHours'):
            parts = ['出差']
            if use_field('travelHours') and r.get('travelHours'):
                parts.append(str(r['travelHours']) + 'h')
            return '/'.join(parts)
        mt = str(r.get('missTime', ''))
        if r.get('missTime') and ('下午' in mt or '下班' in mt or 'PM' in mt):
            return '漏打卡'
        if r.get('status') in ('absent', 'no_sign_in', 'no_sign_out') or r.get('absent'):
            return '缺勤'
        return ''

    dept_employees = {}
    emp_map = {}
    for r in results:
        dept = r.get('department', '未分配')
        dept_employees.setdefault(dept, [])
        eno = r['employeeNo']
        if eno not in dept_employees[dept]:
            dept_employees[dept].append(eno)
        if eno not in emp_map:
            emp_map[eno] = {'name': r.get('name', ''), 'department': dept}

    dept_cols = [{'department': d, 'employees': sorted(emps)} for d, emps in dept_employees.items()]

    emp_date_results = {}
    for r in results:
        key = (r['employeeNo'], r['date'])
        if key not in emp_date_results:
            emp_date_results[key] = r

    wb = Workbook()
    ws = wb.active
    ws.title = f'{y}年{m}月考勤明细'

    header_row1 = ['日期', '排班', '打卡时间']
    header_row2 = ['', '', '']
    for dc in dept_cols:
        for i, eno in enumerate(dc['employees']):
            header_row1.append(dc['department'] if i == 0 else '')
            header_row2.append(emp_map[eno]['name'])

    for c, val in enumerate(header_row1, 1):
        cell = ws.cell(row=1, column=c, value=val)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for c, val in enumerate(header_row2, 1):
        cell = ws.cell(row=2, column=c, value=val)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    dept_start = 4
    for dc in dept_cols:
        if len(dc['employees']) > 1:
            ws.merge_cells(start_row=1, start_column=dept_start, end_row=1, end_column=dept_start + len(dc['employees']) - 1)
        dept_start += len(dc['employees'])

    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')

    current_row = 3

    holiday_map = {}
    if holidays:
        for h in holidays:
            if h.get('date'):
                holiday_map[h['date']] = h

    schedule_by_date = {}
    for sched in schedules:
        work_days = sched.get('workDays', {})
        ws_time = sched.get('workStartTime', '')
        we_time = sched.get('workEndTime', '')
        for day_key, is_work in work_days.items():
            if is_work and ws_time:
                date_key = f'{target_month}-{str(int(day_key)).zfill(2)}'
                schedule_by_date[date_key] = {'workStartTime': ws_time, 'workEndTime': we_time}

    for d in range(1, last_day + 1):
        date_str = f'{target_month}-{str(d).zfill(2)}'
        day_num = str(d).zfill(2)
        date_serial = f'{m}月{d}日'

        holiday = holiday_map.get(date_str)
        if holiday:
            is_rest = not holiday.get('isWorkday', False)
            schedule_label = holiday.get('name', '假期')
        else:
            sched = next((s for s in schedules if s.get('workDays', {}).get(day_num) == True), None)
            is_rest = (sched is None) and len(schedules) > 0
            schedule_label = '休息日' if is_rest else '工作日'

        row_am = current_row
        for c_idx in range(1, 4):
            cell = ws.cell(row=row_am, column=c_idx)
            cell.border = THIN_BORDER
            if is_rest:
                cell.fill = GRAY_FILL

        ws.cell(row=row_am, column=1, value=date_serial)
        ws.cell(row=row_am, column=2, value=schedule_label)
        ws.cell(row=row_am, column=3, value='上午')
        ws.merge_cells(start_row=row_am, start_column=1, end_row=row_am + 1, end_column=1)
        ws.merge_cells(start_row=row_am, start_column=2, end_row=row_am + 1, end_column=2)

        col = 4
        for dc in dept_cols:
            for eno in dc['employees']:
                r = emp_date_results.get((eno, date_str))
                am_val = build_am_cell(r)
                cell = ws.cell(row=row_am, column=col, value=am_val)
                cell.border = THIN_BORDER
                cell.number_format = TEXT_FMT

                font, fill = _get_cell_style(am_val)
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if is_rest:
                    cell.fill = GRAY_FILL
                else:
                    sched_info = schedule_by_date.get(date_str)
                    if am_val and _is_time_val(am_val) and sched_info:
                        sst = sched_info.get('workStartTime', '')
                        ami = _time_to_minutes(am_val)
                        sti = _time_to_minutes(sst)
                        if ami is not None and sti is not None and ami > sti:
                            cell.font = RED_FONT
                col += 1

        row_pm = current_row + 1
        for c_idx in range(1, 4):
            cell = ws.cell(row=row_pm, column=c_idx)
            cell.border = THIN_BORDER
            if is_rest:
                cell.fill = GRAY_FILL

        ws.cell(row=row_pm, column=3, value='下午')

        col = 4
        for dc in dept_cols:
            for eno in dc['employees']:
                r = emp_date_results.get((eno, date_str))
                pm_val = build_pm_cell(r)
                cell = ws.cell(row=row_pm, column=col, value=pm_val)
                cell.border = THIN_BORDER
                cell.number_format = TEXT_FMT

                font, fill = _get_cell_style(pm_val)
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if is_rest:
                    cell.fill = GRAY_FILL
                else:
                    sched_info = schedule_by_date.get(date_str)
                    if pm_val and _is_time_val(pm_val) and sched_info:
                        set_ = sched_info.get('workEndTime', '')
                        pmi = _time_to_minutes(pm_val)
                        edi = _time_to_minutes(set_)
                        if pmi is not None and edi is not None and pmi < edi:
                            cell.font = RED_FONT
                col += 1

        current_row += 2

    _st = startTime
    _et = endTime
    if not _st or not _et:
        for sched in schedules:
            ws_time = sched.get('workStartTime', '')
            we_time = sched.get('workEndTime', '')
            if ws_time and not _st:
                _st = ws_time
            if we_time and not _et:
                _et = we_time
            if _st and _et:
                break
        if not _st or not _is_time_val(_st):
            _st = '08:30'
        if not _et or not _is_time_val(_et):
            _et = '17:30'

    if _st and _et and len(dept_cols) > 0:
        total_cols = 3 + sum(len(dc['employees']) for dc in dept_cols)
        last_col = get_column_letter(total_cols)
        last_data_row = current_row - 1
        st_h, st_m = map(int, str(_st).split(':'))
        et_h, et_m = map(int, str(_et).split(':'))
        late_formula = f'AND(D3<>"", ISNUMBER(D3), D3>TIME({st_h},{st_m},0))'
        early_formula = f'AND(D3<>"", ISNUMBER(D3), D3<TIME({et_h},{et_m},0))'
        ws.conditional_formatting.add(
            f'D3:{last_col}{last_data_row}',
            FormulaRule(formula=[late_formula], font=Font(color='FF0000'))
        )

    for c in range(1, ws.max_column + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 6

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def handle_export_flat(handler, body):
    if Workbook is None:
        handler._send_json(500, message='openpyxl not available')
        return

    records = body.get('records', [])
    template = body.get('template', {'fields': []})
    filename = body.get('filename', 'attendance_export.xlsx')
    start_time = body.get('startTime')
    end_time = body.get('endTime')

    try:
        output = build_flat_report(records, template, filename, start_time, end_time)
        _send_xlsx(handler, output, filename)
    except Exception as e:
        try:
            handler._send_json(500, message=str(e))
        except Exception:
            pass


def handle_export_calendar(handler, body):
    if Workbook is None:
        handler._send_json(500, message='openpyxl not available')
        return

    results = body.get('results', [])
    target_month = body.get('targetMonth', '')
    fields = body.get('fields', [])
    schedules = body.get('schedules', [])
    holidays = body.get('holidays', [])
    start_time = body.get('startTime')
    end_time = body.get('endTime')

    if not target_month:
        handler._send_json(400, message='targetMonth required')
        return

    try:
        output = build_calendar_report(target_month, fields, results, schedules, holidays, start_time, end_time)
        filename = body.get('filename', '考勤明细_' + target_month + '.xlsx')
        _send_xlsx(handler, output, filename)
    except Exception as e:
        try:
            handler._send_json(500, message=str(e))
        except Exception:
            pass
