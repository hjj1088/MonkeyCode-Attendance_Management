"""
考勤导出服务器 - 使用 openpyxl 生成带样式的 XLSX
同时提供静态文件服务和导出 API
"""

import http.server
import json
import io
import os
import re
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


RED_FONT = Font(color='FFFF0000')
BLUE_FONT = Font(color='FF0066CC')
GRAY_FILL = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


def _get_cell_style(val):
    """根据单元格值判断样式"""
    font = None
    fill = None
    
    if not val:
        return font, fill
    
    sv = str(val)
    if re.search(r'请假|出差|加班|补卡', sv):
        font = BLUE_FONT
    elif re.search(r'迟|早', sv):
        font = RED_FONT
    elif re.match(r'^\d{1,2}:\d{2}', sv):
        font = RED_FONT
    
    return font, fill


def build_calendar_report(target_month, fields, results, schedules):
    """构建日历报表 XLSX"""
    y, m = map(int, target_month.split('-'))
    
    from calendar import monthrange
    _, last_day = monthrange(y, m)
    
    field_set = set(f['field'] for f in fields) if fields else set()
    use_field = lambda name: not fields or name in field_set
    
    def build_am_cell(r):
        if not r: return ''
        if r.get('status') == 'rest': return ''
        if r.get('status') == 'leave':
            parts = ['请假']
            if use_field('leaveType') and r.get('leaveType'): parts.append(r['leaveType'])
            if use_field('leaveHours') and r.get('leaveHours') is not None: parts.append(str(r['leaveHours']) + 'h')
            return '/'.join(parts)
        if r.get('status') == 'travel':
            parts = ['出差']
            if use_field('travelHours') and r.get('travelHours'): parts.append(str(r['travelHours']) + 'h')
            return '/'.join(parts)
        if r.get('status') == 'absent' or r.get('absent'): return '缺勤'
        if r.get('status') in ('overtime', 'suspect_ot'):
            parts = ['加班']
            if use_field('overtimeHours') and r.get('overtimeHours'): parts.append(str(r['overtimeHours']) + 'h')
            return '/'.join(parts)
        if use_field('signIn') and r.get('signIn'):
            val = r['signIn']
            if use_field('lateMinutes') and r.get('lateMinutes', 0) > 0:
                val += ' 迟' + str(r['lateMinutes']) + 'min'
            return val
        return ''
    
    def build_pm_cell(r):
        if not r: return ''
        if r.get('status') in ('rest', 'leave', 'travel', 'absent'): return ''
        if r.get('absent'): return ''
        if use_field('signOut') and r.get('signOut'):
            val = r['signOut']
            if use_field('earlyMinutes') and r.get('earlyMinutes', 0) > 0:
                val += ' 早' + str(r['earlyMinutes']) + 'min'
            return val
        return ''
    
    # Group employees by department
    dept_employees = {}
    emp_map = {}
    for r in results:
        dept = r.get('department', '未分配')
        dept_employees.setdefault(dept, [])
        eno = r['employeeNo']
        if eno not in dept_employees[dept]:
            dept_employees[dept].append(eno)
        if eno not in emp_map:
            emp_map[eno] = {'name': r.get('name', ''), 'department': dept}
    
    # Sort departments consistently
    dept_cols = [{'department': d, 'employees': sorted(emps)} for d, emps in dept_employees.items()]
    
    # Group results by employee+date
    emp_date_results = {}
    for r in results:
        key = (r['employeeNo'], r['date'])
        if key not in emp_date_results:
            emp_date_results[key] = r
    
    wb = Workbook()
    ws = wb.active
    ws.title = f'{y}年{m}月考勤明细'
    
    # Row 1: Headers (dates, department headers)
    # Row 2: Employee names
    
    # Column headers
    header_row1 = ['日期', '排班', '打卡时间']
    header_row2 = ['', '', '']
    
    for dc in dept_cols:
        for i, eno in enumerate(dc['employees']):
            header_row1.append(dc['department'] if i == 0 else '')
            header_row2.append(emp_map[eno]['name'])
    
    # Write row 1
    for c, val in enumerate(header_row1, 1):
        cell = ws.cell(row=1, column=c, value=val)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # Write row 2
    for c, val in enumerate(header_row2, 1):
        cell = ws.cell(row=2, column=c, value=val)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # Merge department headers
    dept_start = 4  # Column D onwards (1-indexed)
    for dc in dept_cols:
        if len(dc['employees']) > 1:
            ws.merge_cells(start_row=1, start_column=dept_start, end_row=1, end_column=dept_start + len(dc['employees']) - 1)
        dept_start += len(dc['employees'])
    
    # Merge A1:A2 (日期), B1:B2 (排班), C1:C2 (打卡时间)
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    
    current_row = 3
    
    for d in range(1, last_day + 1):
        date_str = f'{target_month}-{str(d).zfill(2)}'
        day_num = str(d).zfill(2)
        date_serial = f'{m}月{d}日'
        
        sched = next((s for s in schedules if s.get('workDays', {}).get(day_num) == True), None)
        is_rest = (sched is None) and len(schedules) > 0
        schedule_label = '休息日' if is_rest else '工作日'
        
        # AM row
        row_am = current_row
        for c_idx in range(1, 4):
            cell = ws.cell(row=row_am, column=c_idx)
            cell.border = THIN_BORDER
            if is_rest:
                cell.fill = GRAY_FILL
        ws.cell(row=row_am, column=1, value=date_serial)
        ws.cell(row=row_am, column=2, value=schedule_label)
        ws.cell(row=row_am, column=3, value='上午')
        ws.merge_cells(start_row=row_am, start_column=1, end_row=row_am + 1, end_column=1)
        ws.merge_cells(start_row=row_am, start_column=2, end_row=row_am + 1, end_column=2)
        
        col = 4
        for dc in dept_cols:
            for eno in dc['employees']:
                r = emp_date_results.get((eno, date_str))
                am_val = build_am_cell(r)
                cell = ws.cell(row=row_am, column=col, value=am_val)
                cell.border = THIN_BORDER
                
                if is_rest:
                    cell.fill = GRAY_FILL
                else:
                    font, fill = _get_cell_style(am_val)
                    if font: cell.font = font
                    if fill: cell.fill = fill
                col += 1
        
        # PM row
        row_pm = current_row + 1
        for c_idx in range(1, 4):
            cell = ws.cell(row=row_pm, column=c_idx)
            cell.border = THIN_BORDER
            if is_rest:
                cell.fill = GRAY_FILL
        ws.cell(row=row_pm, column=3, value='下午')
        
        col = 4
        for dc in dept_cols:
            for eno in dc['employees']:
                r = emp_date_results.get((eno, date_str))
                pm_val = build_pm_cell(r)
                cell = ws.cell(row=row_pm, column=col, value=pm_val)
                cell.border = THIN_BORDER
                
                if is_rest:
                    cell.fill = GRAY_FILL
                else:
                    font, fill = _get_cell_style(pm_val)
                    if font: cell.font = font
                    if fill: cell.fill = fill
                col += 1
        
        current_row += 2
    
    # Adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_width = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    # Estimate width for CJK characters
                    val_str = str(cell.value)
                    width = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    max_width = max(max_width, width + 2)
        ws.column_dimensions[col_letter].width = min(max_width, 30)
    
    return wb


def build_flat_report(records, template, filename):
    """构建平铺报表 XLSX"""
    wb = Workbook()
    ws = wb.active
    ws.title = '考勤记录'
    
    headers = [f['label'] for f in template['fields']]
    
    # Write header row
    for c, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = Font(bold=True)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # Write data rows
    for i, rec in enumerate(records):
        row_idx = i + 2
        for ci, f in enumerate(template['fields']):
            col_idx = ci + 1
            field_name = f['field']
            
            if field_name == '_index':
                val = i + 1
            else:
                val = rec.get(field_name, '')
            
            val_str = str(val) if val is not None else ''
            cell = ws.cell(row=row_idx, column=col_idx, value=val_str)
            cell.border = THIN_BORDER
            
            font, fill = _get_cell_style(val_str)
            if font: cell.font = font
            if fill: cell.fill = fill
    
    # Adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_width = len(headers[col_idx - 1]) * 2 + 4
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                if cell.value:
                    val_str = str(cell.value)
                    width = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    max_width = max(max_width, width + 2)
        ws.column_dimensions[col_letter].width = min(max_width, 40)
    
    return wb


class ExportHandler(http.server.SimpleHTTPRequestHandler):
    """自定义 HTTP 处理器：静态文件 + API 端点"""
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=os.path.dirname(os.path.abspath(__file__)), **kwargs)
    
    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        
        if parsed.path == '/api/export/flat':
            self._handle_flat_export()
        elif parsed.path == '/api/export/calendar':
            self._handle_calendar_export()
        else:
            self.send_error(404, 'Not Found')
    
    def _handle_flat_export(self):
        try:
            content_len = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_len)
            data = json.loads(body)
            
            records = data.get('records', [])
            template = data.get('template', {'fields': []})
            filename = data.get('filename', 'attendance_export.xlsx')
            
            wb = build_flat_report(records, template, filename)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            from urllib.parse import quote
            safe_filename = quote(filename)
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f"attachment; filename*=UTF-8''{safe_filename}")
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(output.read())
        except Exception as e:
            self.send_error(500, str(e))
    
    def _handle_calendar_export(self):
        try:
            content_len = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_len)
            data = json.loads(body)
            
            target_month = data.get('targetMonth', '')
            fields = data.get('fields', [])
            results = data.get('results', [])
            schedules = data.get('schedules', [])
            
            wb = build_calendar_report(target_month, fields, results, schedules)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f'考勤明细_{target_month}.xlsx'
            
            from urllib.parse import quote
            safe_filename = quote(filename)
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f"attachment; filename*=UTF-8''{safe_filename}")
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(output.read())
        except Exception as e:
            self.send_error(500, str(e))
    
    def do_OPTIONS(self):
        """处理 CORS 预检请求"""
        self.send_response(204)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def log_message(self, format, *args):
        """抑制日志以保持输出清洁"""
        pass


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8000))
    server = http.server.HTTPServer(('0.0.0.0', port), ExportHandler)
    print(f'Server running on port {port}')
    server.serve_forever()
