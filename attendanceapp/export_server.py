"""
考勤导出服务器 - 使用 openpyxl 生成带样式的 XLSX
同时提供静态文件服务和导出 API
"""

import http.server
import json
import io
import os
import platform
import re
import sys
import urllib.parse
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

try:
    from version import APP_VERSION, GIT_SHA, BUILD_TIME        # Overwritten at Docker build time
except ImportError:
    GIT_SHA = 'N/A'
    BUILD_TIME = 'N/A'


RED_FONT = Font(color='FFFF0000')
BLUE_FONT = Font(color='FF0066CC')
GRAY_FILL = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
TEXT_FMT = '@'
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


def _get_cell_style(val):
    font = None
    fill = None
    
    if not val:
        return font, fill
    
    sv = str(val)
    if re.search(r'假|休|出差|加班|补卡', sv):
        font = BLUE_FONT
    elif re.search(r'迟|早|上班未打卡|下班未打卡|缺勤', sv):
        font = RED_FONT
    
    return font, fill


def _is_time_val(v):
    return bool(v and re.match(r'^\d{1,2}:\d{2}$', str(v)))


def _time_to_minutes(t):
    """将 HH:MM 或 H:MM 字符串转换为分钟数，无效返回 None"""
    if not t or not _is_time_val(str(t)):
        return None
    parts = str(t).split(':')
    return int(parts[0]) * 60 + int(parts[1])


def build_calendar_report(target_month, fields, results, schedules, holidays=None, startTime=None, endTime=None):
    """构建日历报表 XLSX"""
    y, m = map(int, target_month.split('-'))
    
    from calendar import monthrange
    _, last_day = monthrange(y, m)
    
    field_set = set(f['field'] for f in fields) if fields else set()
    use_field = lambda name: not fields or name in field_set
    
    def build_am_cell(r):
        if not r: return ''
        if r.get('status') == 'rest': return ''
        if r.get('status') == 'leave':
            parts = []
            if use_field('leaveType') and r.get('leaveType'): parts.append(r['leaveType'])
            if use_field('leaveHours') and r.get('leaveHours') is not None: parts.append(str(r['leaveHours']) + 'h')
            return '/'.join(parts) if parts else '请假'
        if r.get('status') == 'travel':
            parts = ['出差']
            if use_field('travelHours') and r.get('travelHours'): parts.append(str(r['travelHours']) + 'h')
            return '/'.join(parts)
        if r.get('status') == 'absent' or r.get('absent'): return '缺勤'
        if r.get('status') == 'no_sign_in': return '上班未打卡'
        if use_field('signIn') and r.get('signIn'):
            return r['signIn']
        return ''
    
    def build_pm_cell(r):
        if not r: return ''
        if r.get('status') in ('rest', 'leave', 'travel', 'absent'): return ''
        if r.get('absent'): return ''
        if r.get('status') == 'no_sign_out': return '下班未打卡'
        if use_field('signOut') and r.get('signOut'):
            return r['signOut']
        return ''
    
    # Group employees by department
    dept_employees = {}
    emp_map = {}
    for r in results:
        dept = r.get('department', '未分配')
        dept_employees.setdefault(dept, [])
        eno = r['employeeNo']
        if eno not in dept_employees[dept]:
            dept_employees[dept].append(eno)
        if eno not in emp_map:
            emp_map[eno] = {'name': r.get('name', ''), 'department': dept}
    
    # Sort departments consistently
    dept_cols = [{'department': d, 'employees': sorted(emps)} for d, emps in dept_employees.items()]
    
    # Group results by employee+date
    emp_date_results = {}
    for r in results:
        key = (r['employeeNo'], r['date'])
        if key not in emp_date_results:
            emp_date_results[key] = r
    
    wb = Workbook()
    ws = wb.active
    ws.title = f'{y}年{m}月考勤明细'
    
    # Row 1: Headers (dates, department headers)
    # Row 2: Employee names
    
    # Column headers
    header_row1 = ['日期', '排班', '打卡时间']
    header_row2 = ['', '', '']
    
    for dc in dept_cols:
        for i, eno in enumerate(dc['employees']):
            header_row1.append(dc['department'] if i == 0 else '')
            header_row2.append(emp_map[eno]['name'])
    
    # Write row 1
    for c, val in enumerate(header_row1, 1):
        cell = ws.cell(row=1, column=c, value=val)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # Write row 2
    for c, val in enumerate(header_row2, 1):
        cell = ws.cell(row=2, column=c, value=val)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # Merge department headers
    dept_start = 4  # Column D onwards (1-indexed)
    for dc in dept_cols:
        if len(dc['employees']) > 1:
            ws.merge_cells(start_row=1, start_column=dept_start, end_row=1, end_column=dept_start + len(dc['employees']) - 1)
        dept_start += len(dc['employees'])
    
    # Merge A1:A2 (日期), B1:B2 (排班), C1:C2 (打卡时间)
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    
    current_row = 3
    
    holiday_map = {}
    if holidays:
        for h in holidays:
            if h.get('date'):
                holiday_map[h['date']] = h
    
    schedule_by_date = {}
    for sched in schedules:
        work_days = sched.get('workDays', {})
        ws_time = sched.get('workStartTime', '')
        we_time = sched.get('workEndTime', '')
        for day_key, is_work in work_days.items():
            if is_work and ws_time:
                date_key = f'{target_month}-{str(int(day_key)).zfill(2)}'
                schedule_by_date[date_key] = {'workStartTime': ws_time, 'workEndTime': we_time}
    
    for d in range(1, last_day + 1):
        date_str = f'{target_month}-{str(d).zfill(2)}'
        day_num = str(d).zfill(2)
        date_serial = f'{m}月{d}日'
        
        holiday = holiday_map.get(date_str)
        if holiday:
            is_rest = not holiday.get('isWorkday', False)
            schedule_label = holiday.get('name', '假期')
        else:
            sched = next((s for s in schedules if s.get('workDays', {}).get(day_num) == True), None)
            is_rest = (sched is None) and len(schedules) > 0
            schedule_label = '休息日' if is_rest else '工作日'
        
        # AM row
        row_am = current_row
        for c_idx in range(1, 4):
            cell = ws.cell(row=row_am, column=c_idx)
            cell.border = THIN_BORDER
            if is_rest:
                cell.fill = GRAY_FILL
        ws.cell(row=row_am, column=1, value=date_serial)
        ws.cell(row=row_am, column=2, value=schedule_label)
        ws.cell(row=row_am, column=3, value='上午')
        ws.merge_cells(start_row=row_am, start_column=1, end_row=row_am + 1, end_column=1)
        ws.merge_cells(start_row=row_am, start_column=2, end_row=row_am + 1, end_column=2)
        
        col = 4
        for dc in dept_cols:
            for eno in dc['employees']:
                r = emp_date_results.get((eno, date_str))
                am_val = build_am_cell(r)
                cell = ws.cell(row=row_am, column=col, value=am_val)
                cell.border = THIN_BORDER
                cell.number_format = TEXT_FMT
                
                font, fill = _get_cell_style(am_val)
                if font: cell.font = font
                if fill: cell.fill = fill
                if is_rest:
                    cell.fill = GRAY_FILL
                else:
                    sched_info = schedule_by_date.get(date_str)
                    if am_val and _is_time_val(am_val) and sched_info:
                        sst = sched_info.get('workStartTime', '')
                        ami = _time_to_minutes(am_val)
                        sti = _time_to_minutes(sst)
                        if ami is not None and sti is not None and ami > sti:
                            cell.font = RED_FONT
                col += 1
        
        # PM row
        row_pm = current_row + 1
        for c_idx in range(1, 4):
            cell = ws.cell(row=row_pm, column=c_idx)
            cell.border = THIN_BORDER
            if is_rest:
                cell.fill = GRAY_FILL
        ws.cell(row=row_pm, column=3, value='下午')
        
        col = 4
        for dc in dept_cols:
            for eno in dc['employees']:
                r = emp_date_results.get((eno, date_str))
                pm_val = build_pm_cell(r)
                cell = ws.cell(row=row_pm, column=col, value=pm_val)
                cell.border = THIN_BORDER
                cell.number_format = TEXT_FMT
                
                font, fill = _get_cell_style(pm_val)
                if font: cell.font = font
                if fill: cell.fill = fill
                if is_rest:
                    cell.fill = GRAY_FILL
                else:
                    sched_info = schedule_by_date.get(date_str)
                    if pm_val and _is_time_val(pm_val) and sched_info:
                        set_ = sched_info.get('workEndTime', '')
                        pmi = _time_to_minutes(pm_val)
                        edi = _time_to_minutes(set_)
                        if pmi is not None and edi is not None and pmi < edi:
                            cell.font = RED_FONT
                col += 1
        
        current_row += 2
    
    # Apply conditional formatting rules for late/early (visible in Excel)
    _st = startTime
    _et = endTime
    if not _st or not _et:
        for sched in schedules:
            ws_time = sched.get('workStartTime', '')
            we_time = sched.get('workEndTime', '')
            if ws_time and not _st:
                _st = ws_time
            if we_time and not _et:
                _et = we_time
            if _st and _et:
                break
        # Fallback defaults
        if not _st or not _is_time_val(_st):
            _st = '08:30'
        if not _et or not _is_time_val(_et):
            _et = '17:30'
    
    if _st and _et and _is_time_val(_st) and _is_time_val(_et) and current_row > 3:
        data_end_row = current_row - 1
        data_end_col = dept_start - 1
        if data_end_col >= 4:
            data_range = f'D3:{get_column_letter(data_end_col)}{data_end_row}'
            print(f'[DEBUG] Calendar CF APPLIED range={data_range} _st={_st} _et={_et} rows={current_row-1}', flush=True)
            sh, sm = _st.split(':')
            eh, em = _et.split(':')
            
            late_rule = FormulaRule(
                formula=[f'AND(MOD(ROW(),2)=1,ROW()>=3,D3<>"",NOT(ISERROR(TIMEVALUE(D3))),TIMEVALUE(D3)>TIME({int(sh)},{int(sm)},0))'],
                font=RED_FONT
            )
            ws.conditional_formatting.add(data_range, late_rule)
            
            early_rule = FormulaRule(
                formula=[f'AND(MOD(ROW(),2)=0,ROW()>=4,D3<>"",NOT(ISERROR(TIMEVALUE(D3))),TIMEVALUE(D3)<TIME({int(eh)},{int(em)},0))'],
                font=RED_FONT
            )
            ws.conditional_formatting.add(data_range, early_rule)
    else:
        print(f'[DEBUG] Calendar CF SKIPPED _st={repr(_st)} _et={repr(_et)} current_row={current_row} dept_start={dept_start} results={len(results)} schedules={len(schedules)}', flush=True)
    
    # Adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_width = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    # Estimate width for CJK characters
                    val_str = str(cell.value)
                    width = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    max_width = max(max_width, width + 2)
        ws.column_dimensions[col_letter].width = min(max_width, 30)
    
    return wb


def build_flat_report(records, template, filename, startTime=None, endTime=None):
    """构建平铺报表 XLSX"""
    wb = Workbook()
    ws = wb.active
    ws.title = '考勤记录'
    
    headers = [f['label'] for f in template['fields']]
    
    status_labels = {
        'normal': '正常', 'rest': '休息', 'abnormal': '迟到',
        'leave': '请假', 'travel': '出差', 'absent': '缺勤',
        'overtime': '疑似加班', 'suspect_ot': '疑似加班',
        'no_sign_in': '上班未打卡', 'no_sign_out': '下班未打卡',
    }
    
    # Write header row
    for c, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = Font(bold=True)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # Write data rows
    for i, rec in enumerate(records):
        row_idx = i + 2
        for ci, f in enumerate(template['fields']):
            col_idx = ci + 1
            field_name = f['field']
            
            if field_name == '_index':
                val = i + 1
            else:
                val = rec.get(field_name, '')
                if field_name == 'status':
                    val = status_labels.get(val, val)
            
            val_str = str(val) if val is not None else ''
            cell = ws.cell(row=row_idx, column=col_idx, value=val_str)
            cell.border = THIN_BORDER
            cell.number_format = TEXT_FMT
            
            font, fill = _get_cell_style(val_str)
            if font: cell.font = font
            if fill: cell.fill = fill
            
            sst = rec.get('scheduleStart', '') or startTime
            sed = rec.get('scheduleEnd', '') or endTime
            if field_name == 'signIn' and val and _is_time_val(str(val)) and _is_time_val(sst):
                if _time_to_minutes(str(val)) > _time_to_minutes(sst):
                    cell.font = RED_FONT
            elif field_name == 'signOut' and val and _is_time_val(str(val)) and _is_time_val(sed):
                if _time_to_minutes(str(val)) < _time_to_minutes(sed):
                    cell.font = RED_FONT
    
    # Apply conditional formatting rules for late/early
    _st = startTime
    _et = endTime
    if not _st or not _et:
        for rec in records:
            sst = rec.get('scheduleStart', '')
            sed2 = rec.get('scheduleEnd', '')
            if sst and not _st:
                _st = sst
            if sed2 and not _et:
                _et = sed2
            if _st and _et:
                break
        if not _st or not _is_time_val(_st):
            _st = '08:30'
        if not _et or not _is_time_val(_et):
            _et = '17:30'
    
    sh = None; sm = None; eh = None; em = None
    if _st and _et and _is_time_val(_st) and _is_time_val(_et):
        sh, sm = _st.split(':')
        eh, em = _et.split(':')
    
    if sh is not None:
        data_end_row = len(records) + 1
        cf_count = 0
        for ci, f in enumerate(template['fields']):
            col_letter = get_column_letter(ci + 1)
            col_range = f'{col_letter}2:{col_letter}{data_end_row}'
            cell_ref = f'{col_letter}2'
            field = f.get('field', '')
            flabel = f.get('label', '')
            is_si = field == 'signIn' or ('签到' in flabel)
            is_so = field == 'signOut' or ('签退' in flabel)
            if is_si:
                late_rule = FormulaRule(
                    formula=[f'AND({cell_ref}<>"",NOT(ISERROR(TIMEVALUE({cell_ref}))),TIMEVALUE({cell_ref})>TIME({int(sh)},{int(sm)},0))'],
                    font=RED_FONT
                )
                ws.conditional_formatting.add(col_range, late_rule)
                cf_count += 1
            elif is_so:
                early_rule = FormulaRule(
                    formula=[f'AND({cell_ref}<>"",NOT(ISERROR(TIMEVALUE({cell_ref}))),TIMEVALUE({cell_ref})<TIME({int(eh)},{int(em)},0))'],
                    font=RED_FONT
                )
                ws.conditional_formatting.add(col_range, early_rule)
                cf_count += 1
        print(f'[DEBUG] Flat CF APPLIED _st={_st} _et={_et} records={len(records)} rules_added={cf_count}', flush=True)
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_width = len(headers[col_idx - 1]) * 2 + 4
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                if cell.value:
                    val_str = str(cell.value)
                    width = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    max_width = max(max_width, width + 2)
        ws.column_dimensions[col_letter].width = min(max_width, 40)
    
    return wb


class ExportHandler(http.server.SimpleHTTPRequestHandler):
    """自定义 HTTP 处理器：静态文件 + API 端点"""
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=os.path.dirname(os.path.abspath(__file__)), **kwargs)
    
    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == '/health':
            self._handle_health()
        else:
            super().do_GET()
    
    def _handle_health(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        info = {
            'status': 'ok',
            'service': 'attendance-export-server',
            'app_version': APP_VERSION,
            'version': GIT_SHA,
            'build_time': BUILD_TIME,
            'python': platform.python_version(),
            'server_time': datetime.now(timezone.utc).isoformat(),
        }
        self.wfile.write(json.dumps(info, ensure_ascii=False).encode('utf-8'))
    
    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        
        if parsed.path == '/api/export/flat':
            self._handle_flat_export()
        elif parsed.path == '/api/export/calendar':
            self._handle_calendar_export()
        else:
            self.send_error(404, 'Not Found')
    
    def _handle_flat_export(self):
        try:
            content_len = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_len)
            data = json.loads(body)
            
            records = data.get('records', [])
            template = data.get('template', {'fields': []})
            filename = data.get('filename', 'attendance_export.xlsx')
            startTime = data.get('startTime', '')
            endTime = data.get('endTime', '')
            
            wb = build_flat_report(records, template, filename, startTime, endTime)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            from urllib.parse import quote
            safe_filename = quote(filename)
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f"attachment; filename*=UTF-8''{safe_filename}")
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(output.read())
        except Exception as e:
            self.send_error(500, str(e))
    
    def _handle_calendar_export(self):
        try:
            content_len = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_len)
            data = json.loads(body)
            
            target_month = data.get('targetMonth', '')
            fields = data.get('fields', [])
            results = data.get('results', [])
            schedules = data.get('schedules', [])
            holidays = data.get('holidays', [])
            startTime = data.get('startTime', '')
            endTime = data.get('endTime', '')
            
            wb = build_calendar_report(target_month, fields, results, schedules, holidays, startTime, endTime)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f'考勤明细_{target_month}.xlsx'
            
            from urllib.parse import quote
            safe_filename = quote(filename)
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f"attachment; filename*=UTF-8''{safe_filename}")
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(output.read())
        except Exception as e:
            self.send_error(500, str(e))
    
    def do_OPTIONS(self):
        """处理 CORS 预检请求"""
        self.send_response(204)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def log_message(self, fmt, *args):
        msg = "%s - - [%s] %s\n" % (
            self.client_address[0],
            self.log_date_time_string(),
            fmt % args,
        )
        sys.stderr.write(msg)
        sys.stderr.flush()


def _startup_banner(port):
    lines = [
        "=" * 60,
        "  考勤导出服务器  Attendance Export Server",
        "=" * 60,
        f"  Git SHA      : {GIT_SHA}",
        f"  Build Time   : {BUILD_TIME}",
        f"  Python       : {platform.python_version()}",
        f"  Port         : {port}",
        f"  Health Check : http://0.0.0.0:{port}/health",
        f"  Start Time   : {datetime.now(timezone.utc).isoformat()}",
        "=" * 60,
        "",
    ]
    for line in lines:
        print(line, flush=True)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8000))
    _startup_banner(port)
    server = http.server.HTTPServer(('0.0.0.0', port), ExportHandler)
    sys.stdout.flush()
    server.serve_forever()
